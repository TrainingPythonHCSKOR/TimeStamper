####################################################################################################
# Time Stamper (mouse & keyboard hooking)
####################################################################################################
# Author  : Minjae Choi, Minju Shin @HCS
# Version : 0.1.0 -rc
####################################################################################################

# ============================================
# Header
# ============================================
#from __future__ import print_function
import os
import sys
import mouse
import keyboard
import openpyxl
import time
import win32gui
import threading
import re
import TimeStamperUi
import resource_rc
from datetime import datetime, timedelta
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic, QtCore
from PyQt5.QtGui import QIcon

# ============================================
# Global Variables
# ============================================
debug = True            # logging enable or disable (True or False)
delta             = 1   # unit : second 

startTimeStamp    = 0   # start variable
mouseTimeStamp    = 0   # mouse starting variable
keyboardTimeStamp = 0   # keyboard starting variable

preTime     = datetime(2000, 1, 1, 0, 0, 0) # 시간 계산을 위한 preTime
startTime   = datetime(2000, 1, 1, 0, 0, 0) # GUI display initial Time1
endTime     = datetime(2000, 1, 1, 0, 0, 0) # GUI display initial Time2 
workingTime = datetime(2000, 1, 1, 0, 0, 0) # GUI display initial Time3
# ============================================
# UI file load  ui파일을 로드해서 사용(현재는 TimeStamper.py 를 상속받음)
#formClass   = uic.loadUiType("TimeStamper.ui")[0]

# ============================================
# External Definitions
# ============================================
# thread name, log message display
def Debug(*messages):
    if (debug == True):
        print(threading.currentThread().getName(), str(messages))   
#end def:Debug()

# Check time and Caculate the Interval
def TimeCheck():
    global preTime
    nowTime = datetime.now()
    deltaTime = nowTime - preTime
    if (deltaTime.seconds >= delta):
        preTime = nowTime
        return nowTime, True
    else:
        return nowTime, False
# end def:TimeCheck()

# Generate Excel Sheet Format
def MakeTemplit(xlFile,sheetName,sheetExist):
	if sheetExist == 0:
		defaultSheet = xlFile.create_sheet(sheetName)
	else:
		defaultSheet = xlFile.active
		defaultSheet.title = sheetName 
	defaultSheet.append(['DateTime','StartTime','EndTime','WorkTime'])

	for i in range(2,33):
		if i <= 10:
			defaultSheet.cell(row=i,column=1).value = datetime.now().strftime('%b_0')+str(i-1)
		else: 
			defaultSheet.cell(row=i,column=1).value = datetime.now().strftime('%b_')+str(i-1)
		defaultSheet.cell(row=i,column=4).number_format = 'HH:MM:SS'
    # end for	
# end def:MakeTemplit()

def WriteToExcel(now):
    nYear  = str(now.year)
    nMonth = str(now.month)
    nDay   = str(now.day)
    nHour  = str(now.hour)
    nMin   = str(now.minute)
    nSec   = str(now.second)

    global startTime
    global endTime
    global workingTime

    sheetExist = 0
    filePathName = 'Harman_Time_Log.xlsx'
    sheetName = now.strftime('%B')+'_Time_Log'

    if not os.path.exists(filePathName):
        xlFile = openpyxl.Workbook()
        MakeTemplit(xlFile,sheetName,1)
    else:
        xlFile = openpyxl.load_workbook(filePathName)

	# Search sheet name on the file of this month
    for i in range(len(xlFile.sheetnames)):
        if str(xlFile.sheetnames[i]) == sheetName:
            sheetExist = 1
            #break
        else:
            sheetExist = 0
    # end for
	
    if sheetExist:
        sheet = xlFile[sheetName]
    else:
        MakeTemplit(xlFile,sheetName,sheetExist)
        sheet = xlFile[sheetName]

	# Record Log Time 
    xlRowRange = len(sheet[2:32]) 
    # 1일부터 31일 중 해당하는 일자 찾기
    for i in range(2,xlRowRange+2):
        if sheet.cell(row=i,column=1).value == now.strftime('%b_%d'):
            # 시작 시간 없으면 기록
            if not sheet.cell(row=i,column=2).value:
                sheet.cell(row=i,column=2).value = now.strftime('%H:%M:%S')
            # 시작 시간 있으면 마지막 시간 기록
            day_tmp = nDay
            time_tmp = sheet.cell(row=i,column=2).value
            time_tmp = re.split('\:+', time_tmp)
            sheet.cell(row=i,column=3).value = now.strftime('%H:%M:%S')
            sheet.cell(row=i,column=4).value = '=C'+str(i)+'-B'+str(i)
            break
    # end for

    # GUI에 record
    startTime = datetime(int(nYear), int(nMonth), int(day_tmp), int(time_tmp[0]), int(time_tmp[1]), int(time_tmp[2]))
    endTime = datetime(int(nYear), int(nMonth), int(nDay), int(nHour), int(nMin), int(nSec))
    workingTime = endTime - startTime

    xlFile.save(filePathName)
    xlFile.close	
# end : def WriteToExcel()



def OnMouseEvent(event): 
    if(startTimeStamp and mouseTimeStamp):
        Debug("Release Mouse Event")
        now, isTrue = TimeCheck()
        if (isTrue == True):
            Debug("mouse, delta = %d" %delta)
            WriteToExcel(now)
        return True
# end : def OnMouseEvent()

def OnKeyboardEvent(event):
    if(startTimeStamp and keyboardTimeStamp):
        Debug("Release Keyboard Event")
        now, isTrue = TimeCheck()
        if (isTrue == True):
            Debug("keyboard")
            WriteToExcel(now)
        return True
# end : def OnKeyboardEvent()


class MouseKeyboardHandler(threading.Thread):
    def __init__(self):
        super().__init__()
        self.setDaemon(True)
    
    def run(self):
        Debug("is started?")
        mh = threading.Thread(target=mouse.hook(OnMouseEvent))
        kh = threading.Thread(target=keyboard.hook(OnKeyboardEvent))
        win32gui.PumpMessages()
# end class

# UI file 및 기본 UI window 다중 상속 
# UI파일 로드 셋업 및 이벤트컨트롤을 위한 슬롯 연결
#class MyWindow(QMainWindow, formClass):
class MyWindow(QMainWindow,TimeStamperUi.Ui_MainWindow):
    tray_icon = None

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # Main refresh GUI
        timer = QtCore.QTimer(self)
        timer.timeout.connect(self.guiUpdate)
        timer.start(delta*1000)

        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        # Slot Connected
        self.checkBox.stateChanged.connect(self.checkBoxState)
        self.checkBox_2.stateChanged.connect(self.checkBoxState2)
        self.lineEdit.textEdited.connect(self.lineCheck)
        self.pushButton.setCheckable(True)
        self.pushButton.toggled.connect(self.btnToggle)
        self.setWindowIcon(QIcon(':image/mainIcon.png'))

        # TrayIcon
        self.tray_icon = QSystemTrayIcon(self)
        self.tray_icon.setIcon(QIcon(':image/mainIcon.png'))         
        self.tray_icon.activated.connect(self.popupEvent)   
        # TrayIcon rightClick menu    
        quit_action = QAction("Exit", self)                 
        quit_action.triggered.connect(qApp.quit)
        tray_menu = QMenu()
        tray_menu.addAction(quit_action)
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.show()

        MouseKeyboardHandler().start()
    # end def:__init__()
    
    # GUI Time display 
    def guiUpdate(self):
        global startTime
        global endTime
        global workingTime
        self.textEdit.setText(str(startTime))
        self.textEdit_2.setText(str(endTime))
        self.textEdit_3.setText(str(workingTime))
    # end def:guiUpdate()

    # Mouse check handler
    def checkBoxState(self):
        global mouseTimeStamp
        if self.checkBox.isChecked() == True:
            print("Start Detect mouse")
            mouseTimeStamp = 1
        if self.checkBox.isChecked() == False:
            print("Stop Detect mouse")
            mouseTimeStamp = 0
    # end def:checkBoxState()

    # Keyboard check handler
    def checkBoxState2(self):
        global keyboardTimeStamp
        if self.checkBox_2.isChecked() == True:
            print("Start Detect keyboard")
            keyboardTimeStamp = 1
        if self.checkBox_2.isChecked() == False:
            print("Stop Detect keyboard")    
            keyboardTimeStamp = 0
    # end def:checkBoxState2()

    # Unit Seconds handler
    def lineCheck(self):
        global delta
        delta_tmp = self.lineEdit.text()
        if (str.isdecimal(delta_tmp)):
            delta = int(delta_tmp)
    # end def:lineCheck()

    # Start button
    def btnToggle(self):
        global startTimeStamp
        if self.pushButton.isChecked() == True:
            startTimeStamp = 1
            self.pushButton.setText("Running !")
        if self.pushButton.isChecked() == False:
            startTimeStamp = 0
            self.pushButton.setText("Paused...")
    #end def:btnToggle()

    # GUI minimize -> TrayIcon
    def hideEvent(self, event):
        if self.isMinimized:
            self.hide()
    #end def:hideEvent()

    # GUI TrayIcon -> MainWindow 
    def popupEvent(self, reason):
        if reason == QSystemTrayIcon.DoubleClick:
            self.showNormal()
    #end def:popupEvent()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    sys.exit(app.exec_())
